from datetime import date

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Sum
from django.shortcuts import get_object_or_404, redirect, render

from empresa.decorators import requer_admin, requer_empresa
from .forms import DespesaForm, RepasseForm
from .models import CentroCusto, Despesa, FormaPagamento, Repasse, SubGrupo


# ── Cadastros ──────────────────────────────────────────────────────────────────

@login_required
@requer_empresa
@requer_admin
def cadastros(request):
    aba = request.GET.get('aba', 'centros')
    empresa = request.empresa

    if request.method == 'POST':
        acao = request.POST.get('acao')
        nome = request.POST.get('nome', '').strip()
        pk   = request.POST.get('pk')
        model = {'centros': CentroCusto, 'subgrupos': SubGrupo, 'formas': FormaPagamento}.get(aba)

        if model:
            if acao == 'salvar' and nome:
                if pk:
                    obj = get_object_or_404(model, pk=pk, empresa=empresa)
                    obj.nome = nome; obj.save()
                    messages.success(request, 'Cadastro atualizado.')
                else:
                    _, created = model.objects.get_or_create(empresa=empresa, user=request.user, nome=nome)
                    messages.success(request, 'Cadastro adicionado.' if created else 'Já existe um cadastro com esse nome.')
            elif acao == 'excluir' and pk:
                get_object_or_404(model, pk=pk, empresa=empresa).delete()
                messages.success(request, 'Cadastro removido.')

        return redirect(f'{request.path}?aba={aba}')

    return render(request, 'financeiro/cadastros.html', {
        'aba': aba,
        'panels': [
            ('centros',   'Centro de Custo',   CentroCusto.objects.filter(empresa=empresa)),
            ('subgrupos', 'SubGrupo',           SubGrupo.objects.filter(empresa=empresa)),
            ('formas',    'Forma de Pagamento', FormaPagamento.objects.filter(empresa=empresa)),
        ],
    })


# ── Despesas ───────────────────────────────────────────────────────────────────

@login_required
@requer_empresa
def despesa_list(request):
    qs = Despesa.objects.filter(empresa=request.empresa)
    hoje = date.today()
    de       = request.GET.get('de', hoje.replace(month=hoje.month-1 if hoje.month > 1 else 12, year=hoje.year if hoje.month > 1 else hoje.year-1, day=1).isoformat())
    ate      = request.GET.get('ate', hoje.isoformat())
    centro   = request.GET.get('centro', '').strip()
    subgrupo = request.GET.get('subgrupo', '').strip()
    situacao = request.GET.get('situacao', '')

    try:
        from datetime import datetime
        qs = qs.filter(data__range=(datetime.strptime(de, '%Y-%m-%d').date(),
                                    datetime.strptime(ate, '%Y-%m-%d').date()))
    except ValueError:
        pass
    if centro:
        qs = qs.filter(centro_custo=centro)
    if subgrupo:
        qs = qs.filter(subgrupo=subgrupo)
    if situacao:
        qs = qs.filter(situacao=situacao)

    total    = qs.aggregate(total=Sum('valor'))['total'] or 0
    paginator = Paginator(qs, 20)
    page     = paginator.get_page(request.GET.get('page'))

    centros   = CentroCusto.objects.filter(empresa=request.empresa).values_list('nome', flat=True)
    subgrupos = SubGrupo.objects.filter(empresa=request.empresa).values_list('nome', flat=True)

    return render(request, 'financeiro/despesa_list.html', {
        'page_obj': page, 'total': total,
        'de': de, 'ate': ate, 'centro': centro, 'subgrupo': subgrupo, 'situacao': situacao,
        'centros': centros, 'subgrupos': subgrupos, 'situacoes': Despesa.Situacao,
    })


@login_required
@requer_empresa
def despesa_create(request):
    form = DespesaForm(request.POST or None, empresa=request.empresa)
    if form.is_valid():
        d = form.save(commit=False)
        d.user = request.user
        d.empresa = request.empresa
        d.save()
        messages.success(request, 'Despesa cadastrada com sucesso.')
        return redirect('despesa_list')
    return render(request, 'financeiro/despesa_form.html', {'form': form, 'titulo': 'Nova Despesa'})


@login_required
@requer_empresa
@requer_admin
def despesa_edit(request, pk):
    despesa = get_object_or_404(Despesa, pk=pk, empresa=request.empresa)
    form = DespesaForm(request.POST or None, instance=despesa, empresa=request.empresa)
    if form.is_valid():
        form.save()
        messages.success(request, 'Despesa atualizada.')
        return redirect('despesa_list')
    return render(request, 'financeiro/despesa_form.html', {'form': form, 'titulo': 'Editar Despesa'})


@login_required
@requer_empresa
@requer_admin
def despesa_delete(request, pk):
    despesa = get_object_or_404(Despesa, pk=pk, empresa=request.empresa)
    if request.method == 'POST':
        despesa.delete()
        messages.success(request, 'Despesa excluída.')
        return redirect('despesa_list')
    return render(request, 'financeiro/despesa_confirm_delete.html', {'despesa': despesa})


@login_required
@requer_empresa
@requer_admin
def despesa_bulk_delete(request):
    if request.method != 'POST':
        return redirect('despesa_list')
    ids = request.POST.getlist('ids')
    if ids:
        deleted, _ = Despesa.objects.filter(pk__in=ids, empresa=request.empresa).delete()
        messages.success(request, f'{deleted} despesa(s) excluída(s).')
    else:
        messages.info(request, 'Nenhuma despesa selecionada.')
    return redirect('despesa_list')


# ── Repasses ───────────────────────────────────────────────────────────────────

@login_required
@requer_empresa
def repasse_list(request):
    qs      = Repasse.objects.filter(empresa=request.empresa)
    hoje = date.today()
    de      = request.GET.get('de', hoje.replace(month=hoje.month-1 if hoje.month > 1 else 12, year=hoje.year if hoje.month > 1 else hoje.year-1, day=1).isoformat())
    ate     = request.GET.get('ate', hoje.isoformat())
    tipo    = request.GET.get('tipo', '')
    origem  = request.GET.get('origem', '').strip()
    destino = request.GET.get('destino', '').strip()

    try:
        from datetime import datetime
        qs = qs.filter(data__range=(datetime.strptime(de, '%Y-%m-%d').date(),
                                    datetime.strptime(ate, '%Y-%m-%d').date()))
    except ValueError:
        pass
    if tipo:
        qs = qs.filter(tipo=tipo)
    if origem:
        qs = qs.filter(origem=origem)
    if destino:
        qs = qs.filter(destino=destino)

    origens  = Repasse.objects.filter(empresa=request.empresa).values_list('origem', flat=True).distinct().order_by('origem')
    destinos = Repasse.objects.filter(empresa=request.empresa).values_list('destino', flat=True).distinct().order_by('destino')

    paginator = Paginator(qs, 20)
    page = paginator.get_page(request.GET.get('page'))
    return render(request, 'financeiro/repasse_list.html', {
        'page_obj': page, 'de': de, 'ate': ate, 'tipo': tipo, 'origem': origem,
        'destino': destino, 'tipos': Repasse.Tipo,
        'origens': origens, 'destinos': destinos,
    })


@login_required
@requer_empresa
def repasse_create(request):
    form = RepasseForm(request.POST or None, empresa=request.empresa)
    if form.is_valid():
        r = form.save(commit=False)
        r.user = request.user
        r.empresa = request.empresa
        r.save()
        messages.success(request, 'Repasse cadastrado com sucesso.')
        return redirect('repasse_list')
    return render(request, 'financeiro/repasse_form.html', {'form': form, 'titulo': 'Novo Repasse'})


@login_required
@requer_empresa
@requer_admin
def repasse_edit(request, pk):
    repasse = get_object_or_404(Repasse, pk=pk, empresa=request.empresa)
    form = RepasseForm(request.POST or None, instance=repasse, empresa=request.empresa)
    if form.is_valid():
        form.save()
        messages.success(request, 'Repasse atualizado.')
        return redirect('repasse_list')
    return render(request, 'financeiro/repasse_form.html', {'form': form, 'titulo': 'Editar Repasse'})


@login_required
@requer_empresa
@requer_admin
def repasse_delete(request, pk):
    repasse = get_object_or_404(Repasse, pk=pk, empresa=request.empresa)
    if request.method == 'POST':
        repasse.delete()
        messages.success(request, 'Repasse excluído.')
        return redirect('repasse_list')
    return render(request, 'financeiro/repasse_confirm_delete.html', {'repasse': repasse})


@login_required
@requer_empresa
@requer_admin
def repasse_bulk_delete(request):
    if request.method != 'POST':
        return redirect('repasse_list')
    ids = request.POST.getlist('ids')
    if ids:
        deleted, _ = Repasse.objects.filter(pk__in=ids, empresa=request.empresa).delete()
        messages.success(request, f'{deleted} repasse(s) excluído(s).')
    else:
        messages.info(request, 'Nenhum repasse selecionado.')
    return redirect('repasse_list')


# ── Dashboard ──────────────────────────────────────────────────────────────────

@login_required
@requer_empresa
def financeiro_dashboard(request):
    mes_param = request.GET.get('mes', '')
    try:
        ano, mes = mes_param.split('-')
        ano, mes = int(ano), int(mes)
    except ValueError:
        hoje = date.today()
        ano, mes = hoje.year, hoje.month

    import calendar
    mes_label = f"{calendar.month_name[mes].capitalize()}/{ano}"
    mes_val   = f'{ano:04d}-{mes:02d}'

    empresa = request.empresa
    repasses_mes = Repasse.objects.filter(empresa=empresa, data__year=ano, data__month=mes)
    despesas_mes = Despesa.objects.filter(empresa=empresa, data__year=ano, data__month=mes)

    total_aportes  = repasses_mes.filter(tipo='aporte').aggregate(v=Sum('valor'))['v'] or 0
    total_repasses = repasses_mes.filter(tipo='repasse').aggregate(v=Sum('valor'))['v'] or 0
    total_despesas = despesas_mes.aggregate(v=Sum('valor'))['v'] or 0
    saldo          = total_aportes - total_repasses - total_despesas

    por_centro   = list(despesas_mes.values('centro_custo').annotate(total=Sum('valor')).order_by('-total'))
    por_subgrupo = list(despesas_mes.values('subgrupo').annotate(total=Sum('valor')).order_by('-total'))

    from itertools import chain
    import operator
    movs_r = [{'data': r.data, 'tipo': r.get_tipo_display(), 'descricao': f'{r.origem} → {r.destino}', 'valor': r.valor, 'badge': 'blue'}
              for r in Repasse.objects.filter(empresa=empresa).order_by('-data', '-created_at')[:20]]
    movs_d = [{'data': d.data, 'tipo': 'Despesa', 'descricao': d.descricao, 'valor': d.valor, 'badge': 'red'}
              for d in Despesa.objects.filter(empresa=empresa).order_by('-data', '-created_at')[:20]]
    ultimas = sorted(chain(movs_r, movs_d), key=operator.itemgetter('data'), reverse=True)[:10]

    return render(request, 'financeiro/dashboard.html', {
        'total_aportes': total_aportes, 'total_repasses': total_repasses,
        'total_despesas': total_despesas, 'saldo': saldo,
        'por_centro': por_centro, 'por_subgrupo': por_subgrupo,
        'ultimas': ultimas, 'mes_label': mes_label, 'mes_val': mes_val,
    })


# ── Relatórios ─────────────────────────────────────────────────────────────────

def _get_periodo(request):
    mes_param = request.GET.get('mes', '')
    try:
        ano, mes = mes_param.split('-')
        ano, mes = int(ano), int(mes)
    except ValueError:
        hoje = date.today()
        ano, mes = hoje.year, hoje.month
    import calendar
    return ano, mes, f"{calendar.month_name[mes].capitalize()}/{ano}"


@login_required
@requer_empresa
def relatorios(request):
    ano, mes, label = _get_periodo(request)
    return render(request, 'financeiro/relatorios.html', {
        'mes_label': label, 'mes_val': f'{ano:04d}-{mes:02d}',
        'tipo': request.GET.get('tipo', ''), 'fmt': request.GET.get('fmt', 'pdf'),
    })


def _pdf_response(filename):
    from django.http import HttpResponse
    r = HttpResponse(content_type='application/pdf')
    r['Content-Disposition'] = f'attachment; filename="{filename}"'
    return r


def _build_pdf(title, headers, rows, totais=None):
    import io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    data   = [headers] + rows + ([totais] if totais else [])
    page_w = landscape(A4)[0] - 3*cm
    t = Table(data, colWidths=[page_w / len(headers)] * len(headers), repeatRows=1)
    style = [
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e1b4b')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 9),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
        ('GRID',       (0,0), (-1,-1), 0.3, colors.HexColor('#e2e8f0')),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 5), ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]
    if totais:
        style += [('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#e0e7ff')),
                  ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold')]
    t.setStyle(TableStyle(style))
    doc.build([Paragraph(title, styles['Title']), Spacer(1, .4*cm), t])
    return buf.getvalue()


def _excel_response(filename):
    from django.http import HttpResponse
    r = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    r['Content-Disposition'] = f'attachment; filename="{filename}"'
    return r


def _build_excel(title, headers, rows, totais=None):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook(); ws = wb.active
    ws.title = title[:31].replace('/', '-')
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws['A1'].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(headers)
    hfill = PatternFill('solid', fgColor='1e1b4b')
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color='FFFFFF'); cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')
    for i, row in enumerate(rows):
        ws.append(row)
        if i % 2 == 1:
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill('solid', fgColor='F8FAFC')
    if totais:
        ws.append(totais)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True); cell.fill = PatternFill('solid', fgColor='E0E7FF')
    for col in ws.columns:
        try:
            ml = max((len(str(c.value or '')) for c in col if hasattr(c, 'column_letter')), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(ml + 4, 40)
        except (AttributeError, TypeError):
            pass
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


@login_required
@requer_empresa
def rel_despesas_centro_pdf(request):
    ano, mes, label = _get_periodo(request)
    qs = (Despesa.objects.filter(empresa=request.empresa, data__year=ano, data__month=mes)
          .values('centro_custo').annotate(total=Sum('valor')).order_by('centro_custo'))
    rows = [[r['centro_custo'], f"R$ {r['total']}"] for r in qs]
    grand = sum(r['total'] for r in qs)
    pdf = _build_pdf(f'Despesas por Centro — {label}', ['Centro de Custo', 'Total'], rows, ['TOTAL', f'R$ {grand}'])
    resp = _pdf_response(f'despesas_centro_{ano}_{mes:02d}.pdf'); resp.write(pdf); return resp


@login_required
@requer_empresa
def rel_despesas_centro_excel(request):
    ano, mes, label = _get_periodo(request)
    qs = (Despesa.objects.filter(empresa=request.empresa, data__year=ano, data__month=mes)
          .values('centro_custo').annotate(total=Sum('valor')).order_by('centro_custo'))
    rows = [[r['centro_custo'], float(r['total'])] for r in qs]
    grand = sum(r['total'] for r in qs)
    data = _build_excel(f'Despesas por Centro — {label}', ['Centro de Custo', 'Total (R$)'], rows, ['TOTAL', float(grand)])
    resp = _excel_response(f'despesas_centro_{ano}_{mes:02d}.xlsx'); resp.write(data); return resp


@login_required
@requer_empresa
def rel_despesas_subgrupo_pdf(request):
    ano, mes, label = _get_periodo(request)
    qs = (Despesa.objects.filter(empresa=request.empresa, data__year=ano, data__month=mes)
          .values('subgrupo').annotate(total=Sum('valor')).order_by('subgrupo'))
    rows = [[r['subgrupo'], f"R$ {r['total']}"] for r in qs]
    grand = sum(r['total'] for r in qs)
    pdf = _build_pdf(f'Despesas por SubGrupo — {label}', ['SubGrupo', 'Total'], rows, ['TOTAL', f'R$ {grand}'])
    resp = _pdf_response(f'despesas_subgrupo_{ano}_{mes:02d}.pdf'); resp.write(pdf); return resp


@login_required
@requer_empresa
def rel_despesas_subgrupo_excel(request):
    ano, mes, label = _get_periodo(request)
    qs = (Despesa.objects.filter(empresa=request.empresa, data__year=ano, data__month=mes)
          .values('subgrupo').annotate(total=Sum('valor')).order_by('subgrupo'))
    rows = [[r['subgrupo'], float(r['total'])] for r in qs]
    grand = sum(r['total'] for r in qs)
    data = _build_excel(f'Despesas por SubGrupo — {label}', ['SubGrupo', 'Total (R$)'], rows, ['TOTAL', float(grand)])
    resp = _excel_response(f'despesas_subgrupo_{ano}_{mes:02d}.xlsx'); resp.write(data); return resp


@login_required
@requer_empresa
def rel_repasses_pdf(request):
    ano, mes, label = _get_periodo(request)
    qs = Repasse.objects.filter(empresa=request.empresa, data__year=ano, data__month=mes).order_by('data')
    rows = [[str(r.data), r.get_tipo_display(), r.origem, r.destino, f'R$ {r.valor}', r.descricao] for r in qs]
    grand = qs.aggregate(v=Sum('valor'))['v'] or 0
    pdf = _build_pdf(f'Repasses — {label}', ['Data','Tipo','Origem','Destino','Valor','Descrição'], rows, ['','','','TOTAL',f'R$ {grand}',''])
    resp = _pdf_response(f'repasses_{ano}_{mes:02d}.pdf'); resp.write(pdf); return resp


@login_required
@requer_empresa
def rel_repasses_excel(request):
    ano, mes, label = _get_periodo(request)
    qs = Repasse.objects.filter(empresa=request.empresa, data__year=ano, data__month=mes).order_by('data')
    rows = [[str(r.data), r.get_tipo_display(), r.origem, r.destino, float(r.valor), r.descricao] for r in qs]
    grand = qs.aggregate(v=Sum('valor'))['v'] or 0
    data = _build_excel(f'Repasses — {label}', ['Data','Tipo','Origem','Destino','Valor','Descrição'], rows, ['','','','TOTAL',float(grand),''])
    resp = _excel_response(f'repasses_{ano}_{mes:02d}.xlsx'); resp.write(data); return resp


@login_required
@requer_empresa
def rel_completo_pdf(request):
    import io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    ano, mes, label = _get_periodo(request)
    styles  = getSampleStyleSheet()
    page_w  = landscape(A4)[0] - 3*cm

    def make_table(headers, rows, totais=None):
        data = [headers] + rows + ([totais] if totais else [])
        t = Table(data, colWidths=[page_w / len(headers)] * len(headers), repeatRows=1)
        s = [('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1e1b4b')),
             ('TEXTCOLOR',(0,0),(-1,0),colors.white),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
             ('FONTSIZE',(0,0),(-1,-1),8),
             ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#f8fafc')]),
             ('GRID',(0,0),(-1,-1),.3,colors.HexColor('#e2e8f0')),
             ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4)]
        if totais:
            s += [('BACKGROUND',(0,-1),(-1,-1),colors.HexColor('#e0e7ff')),
                  ('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold')]
        t.setStyle(TableStyle(s)); return t

    empresa  = request.empresa
    despesas = Despesa.objects.filter(empresa=empresa, data__year=ano, data__month=mes).order_by('data')
    repasses = Repasse.objects.filter(empresa=empresa, data__year=ano, data__month=mes).order_by('data')
    grand_d  = despesas.aggregate(v=Sum('valor'))['v'] or 0
    grand_r  = repasses.aggregate(v=Sum('valor'))['v'] or 0

    rows_d = [[str(d.data),d.centro_custo,d.subgrupo,d.descricao,d.forma_pagamento,d.get_situacao_display(),f'R$ {d.valor}'] for d in despesas]
    rows_r = [[str(r.data),r.get_tipo_display(),r.origem,r.destino,f'R$ {r.valor}',r.descricao] for r in repasses]

    elements = [Paragraph(f'Relatório Completo — {label}', styles['Title']), Spacer(1,.5*cm),
                Paragraph('Despesas', styles['Heading2']), Spacer(1,.2*cm),
                make_table(['Data','Centro','SubGrupo','Descrição','Pagamento','Situação','Valor'], rows_d,
                           ['','','','','','TOTAL',f'R$ {grand_d}']),
                Spacer(1,.6*cm),
                Paragraph('Repasses / Aportes', styles['Heading2']), Spacer(1,.2*cm),
                make_table(['Data','Tipo','Origem','Destino','Valor','Descrição'], rows_r,
                           ['','','','TOTAL',f'R$ {grand_r}',''])]

    buf = io.BytesIO()
    SimpleDocTemplate(buf, pagesize=landscape(A4),
                      leftMargin=1.5*cm, rightMargin=1.5*cm,
                      topMargin=1.5*cm, bottomMargin=1.5*cm).build(elements)
    resp = _pdf_response(f'relatorio_completo_{ano}_{mes:02d}.pdf')
    resp.write(buf.getvalue()); return resp
