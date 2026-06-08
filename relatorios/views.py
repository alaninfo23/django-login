import io
from datetime import date, datetime

from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import render

from assets.models import Asset
from empresa.decorators import requer_empresa
from financeiro.models import CentroCusto, Despesa, FormaPagamento, Repasse, SubGrupo
from frota.models import Abastecimento, Manutencao, Motorista, Veiculo, Viagem


# ── helpers ────────────────────────────────────────────────────────────────────

def _fmt_valor(v):
    """Formata Decimal/float para R$ 1.234,56"""
    return 'R$ {:,.2f}'.format(v).replace(',', 'X').replace('.', ',').replace('X', '.')


def _fmt_data(d):
    """Formata date/str para dd/mm/yyyy"""
    if not d:
        return ''
    if hasattr(d, 'strftime'):
        return d.strftime('%d/%m/%Y')
    try:
        return datetime.strptime(str(d), '%Y-%m-%d').strftime('%d/%m/%Y')
    except ValueError:
        return str(d)


def _parse_periodo(request):
    de_str  = request.GET.get('de', '')
    ate_str = request.GET.get('ate', '')
    try:
        de  = datetime.strptime(de_str,  '%Y-%m-%d').date()
        ate = datetime.strptime(ate_str, '%Y-%m-%d').date()
    except ValueError:
        hoje = date.today()
        de   = hoje.replace(day=1)
        ate  = hoje
    return de, ate


def _pdf_resp(filename):
    r = HttpResponse(content_type='application/pdf')
    r['Content-Disposition'] = f'attachment; filename="{filename}"'
    return r


def _excel_resp(filename):
    r = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    r['Content-Disposition'] = f'attachment; filename="{filename}"'
    return r


def _build_pdf(title, headers, rows, totais=None, col_widths=None, logo_path=None):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    data   = [headers] + rows + ([totais] if totais else [])
    page_w = landscape(A4)[0] - 3*cm
    if col_widths:
        total_parts = sum(col_widths)
        widths = [page_w * (p / total_parts) for p in col_widths]
    else:
        widths = [page_w / len(headers)] * len(headers)
    t = Table(data, colWidths=widths, repeatRows=1)
    style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e1b4b')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1,-1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
        ('GRID',       (0, 0), (-1,-1), 0.3, colors.HexColor('#e2e8f0')),
        ('VALIGN',     (0, 0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1,-1), 5),
        ('BOTTOMPADDING', (0, 0), (-1,-1), 5),
    ]
    if totais:
        style += [('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#e0e7ff')),
                  ('FONTNAME',   (0,-1), (-1,-1), 'Helvetica-Bold')]
    t.setStyle(TableStyle(style))

    story = []
    if logo_path:
        try:
            img = Image(logo_path, height=1.2*cm, width=3.5*cm, kind='proportional')
            story.append(img)
            story.append(Spacer(1, .2*cm))
        except Exception:
            pass
    story += [Paragraph(title, styles['Title']), Spacer(1, .4*cm), t]
    doc.build(story)
    return buf.getvalue()


def _build_excel(title, headers, rows, totais=None, min_widths=None):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31].replace('/', '-')
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws['A1'].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(headers)
    hfill = PatternFill('solid', fgColor='1e1b4b')
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')
    for i, row in enumerate(rows):
        ws.append(row)
        if i % 2 == 1:
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill('solid', fgColor='F8FAFC')
    if totais:
        ws.append(totais)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='E0E7FF')
    for i, col in enumerate(ws.columns):
        try:
            from openpyxl.utils import get_column_letter
            letter = get_column_letter(i + 1)
            vals = []
            for c in col:
                try:
                    vals.append(len(str(c.value or '')))
                except Exception:
                    pass
            ml = max(vals) if vals else 10
            w = min(ml + 4, 60)
            if min_widths and i < len(min_widths):
                w = max(w, min_widths[i])
            ws.column_dimensions[letter].width = w
        except Exception:
            pass
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── central view ───────────────────────────────────────────────────────────────

@login_required
@requer_empresa
def central(request):
    hoje = date.today()
    # padrão: do 1º dia do mês anterior até hoje
    if hoje.month == 1:
        de_default = hoje.replace(year=hoje.year - 1, month=12, day=1)
    else:
        de_default = hoje.replace(month=hoje.month - 1, day=1)
    de  = request.GET.get('de',  de_default.isoformat())
    ate = request.GET.get('ate', hoje.isoformat())
    modulo = request.GET.get('modulo', 'patrimonial')

    empresa = request.empresa
    centros   = CentroCusto.objects.filter(empresa=empresa).values_list('nome', flat=True)
    subgrupos = SubGrupo.objects.filter(empresa=empresa).values_list('nome', flat=True)
    formas    = FormaPagamento.objects.filter(empresa=empresa).values_list('nome', flat=True)
    veiculos  = Veiculo.objects.filter(empresa=empresa).order_by('placa')

    return render(request, 'relatorios/central.html', {
        'de': de, 'ate': ate, 'modulo': modulo,
        'asset_types': Asset.AssetType.choices,
        'asset_statuses': Asset.Status.choices,
        'centros': centros, 'subgrupos': subgrupos, 'formas': formas,
        'veiculos': veiculos,
        'frota_subtipos': [
            ('veiculos',       'Veículos'),
            ('abastecimentos', 'Abastecimentos'),
            ('manutencoes',    'Manutenções'),
            ('viagens',        'Viagens'),
            ('todos',          'Todos'),
        ],
    })


# ── PATRIMONIAL ────────────────────────────────────────────────────────────────

@login_required
@requer_empresa
def patrimonial_pdf(request):
    de, ate = _parse_periodo(request)
    subtipo  = request.GET.get('subtipo', '')
    status   = request.GET.get('status', '')

    qs = Asset.objects.filter(empresa=request.empresa, purchase_date__range=(de, ate))
    if subtipo:
        qs = qs.filter(asset_type=subtipo)
    if status:
        qs = qs.filter(status=status)

    if not qs.exists():
        from django.contrib import messages
        from django.shortcuts import redirect
        messages.warning(request, 'Nenhum dado encontrado para os filtros selecionados.')
        return redirect(request.META.get('HTTP_REFERER', '/relatorios/'))

    headers = ['Nome', 'Tipo', 'Valor (R$)', 'Data Compra', 'Status', 'Local']
    rows    = [[a.name, a.get_asset_type_display(), _fmt_valor(a.acquisition_value),
                _fmt_data(a.purchase_date), a.get_status_display(), a.location] for a in qs]
    total   = qs.aggregate(v=Sum('acquisition_value'))['v'] or 0
    totais  = ['TOTAL', '', _fmt_valor(total), '', '', '']

    label = f'{de.strftime("%d/%m/%Y")} a {ate.strftime("%d/%m/%Y")}'
    logo  = request.empresa.logo.path if request.empresa.logo else None
    pdf = _build_pdf(f'Relatório Patrimonial — {label}', headers, rows, totais,
                     col_widths=[3, 2, 2, 1.5, 1.5, 2], logo_path=logo)
    resp = _pdf_resp(f'patrimonial_{de}_{ate}.pdf')
    resp.write(pdf)
    return resp


@login_required
@requer_empresa
def patrimonial_excel(request):
    de, ate = _parse_periodo(request)
    subtipo  = request.GET.get('subtipo', '')
    status   = request.GET.get('status', '')

    qs = Asset.objects.filter(empresa=request.empresa, purchase_date__range=(de, ate))
    if subtipo:
        qs = qs.filter(asset_type=subtipo)
    if status:
        qs = qs.filter(status=status)

    if not qs.exists():
        from django.contrib import messages
        from django.shortcuts import redirect
        messages.warning(request, 'Nenhum dado encontrado para os filtros selecionados.')
        return redirect(request.META.get('HTTP_REFERER', '/relatorios/'))

    headers = ['Nome', 'Tipo', 'Valor (R$)', 'Data Compra', 'Status', 'Local']
    rows    = [[a.name, a.get_asset_type_display(), _fmt_valor(a.acquisition_value),
                _fmt_data(a.purchase_date), a.get_status_display(), a.location] for a in qs]
    total   = qs.aggregate(v=Sum('acquisition_value'))['v'] or 0
    totais  = ['TOTAL', '', _fmt_valor(total), '', '', '']

    label = f'{de.strftime("%d/%m/%Y")} a {ate.strftime("%d/%m/%Y")}'
    # Nome, Tipo, Valor, Data, Status, Local
    data = _build_excel(f'Patrimonial — {label}', headers, rows, totais,
                        min_widths=[35, 22, 16, 12, 12, 20])
    resp = _excel_resp(f'patrimonial_{de}_{ate}.xlsx')
    resp.write(data)
    return resp


# ── FINANCEIRO ─────────────────────────────────────────────────────────────────

@login_required
@requer_empresa
def financeiro_pdf(request):
    de, ate  = _parse_periodo(request)
    subtipo  = request.GET.get('subtipo', 'despesas')  # despesas | repasses | ambos
    centro   = request.GET.get('centro', '')
    subgrupo = request.GET.get('subgrupo', '')
    forma    = request.GET.get('forma', '')

    label = f'{de.strftime("%d/%m/%Y")} a {ate.strftime("%d/%m/%Y")}'

    import io as _io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buf    = _io.BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=landscape(A4),
                               leftMargin=1.5*cm, rightMargin=1.5*cm,
                               topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    page_w = landscape(A4)[0] - 3*cm
    story  = []
    logo_path = request.empresa.logo.path if request.empresa.logo else None
    if logo_path:
        from reportlab.platypus import Image as RLImage
        try:
            story.append(RLImage(logo_path, height=1.2*cm, width=3.5*cm, kind='proportional'))
            story.append(Spacer(1, .2*cm))
        except Exception:
            pass
    story += [Paragraph(f'Relatório Financeiro — {label}', styles['Title']), Spacer(1, .4*cm)]

    def _tbl(headers, rows, totais=None, col_widths=None):
        data = [headers] + rows + ([totais] if totais else [])
        if col_widths:
            total_parts = sum(col_widths)
            widths = [page_w * (p / total_parts) for p in col_widths]
        else:
            widths = [page_w / len(headers)] * len(headers)
        t = Table(data, colWidths=widths, repeatRows=1)
        s = [('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1e1b4b')),
             ('TEXTCOLOR',(0,0),(-1,0),colors.white),
             ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
             ('FONTSIZE',(0,0),(-1,-1),8),
             ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#f8fafc')]),
             ('GRID',(0,0),(-1,-1),.3,colors.HexColor('#e2e8f0')),
             ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4)]
        if totais:
            s += [('BACKGROUND',(0,-1),(-1,-1),colors.HexColor('#e0e7ff')),
                  ('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold')]
        t.setStyle(TableStyle(s))
        return t

    has_data = False
    if subtipo in ('despesas', 'ambos'):
        qs = Despesa.objects.filter(empresa=request.empresa, data__range=(de, ate))
        if centro:   qs = qs.filter(centro_custo=centro)
        if subgrupo: qs = qs.filter(subgrupo=subgrupo)
        if forma:    qs = qs.filter(forma_pagamento=forma)
        if qs.exists():
            has_data = True
            rows = [[_fmt_data(d.data), d.centro_custo, d.subgrupo, d.descricao,
                     d.forma_pagamento, d.get_situacao_display(), _fmt_valor(d.valor)] for d in qs]
            total = qs.aggregate(v=Sum('valor'))['v'] or 0
            # Data=1, Centro=2.5, SubGrupo=2.5, Descrição=4.5, FormaPag=1.5, Situação=1, Valor=1
            story += [Paragraph('Despesas', styles['Heading2']), Spacer(1, .2*cm),
                      _tbl(['Data','Centro','SubGrupo','Descrição','Forma Pag.','Situação','Valor'],
                           rows, ['','','','','','TOTAL', _fmt_valor(total)],
                           col_widths=[1, 2.5, 2.5, 4.5, 1.5, 1, 1]),
                      Spacer(1, .5*cm)]

    if subtipo in ('repasses', 'ambos'):
        qs = Repasse.objects.filter(empresa=request.empresa, data__range=(de, ate))
        if qs.exists():
            has_data = True
            rows = [[_fmt_data(r.data), r.get_tipo_display(), r.origem, r.destino,
                     _fmt_valor(r.valor), r.descricao] for r in qs]
            total = qs.aggregate(v=Sum('valor'))['v'] or 0
            story += [Paragraph('Repasses', styles['Heading2']), Spacer(1, .2*cm),
                      _tbl(['Data','Tipo','Origem','Destino','Valor','Descrição'],
                           rows, ['','','','TOTAL', _fmt_valor(total), ''],
                           col_widths=[1, 1.5, 2, 2, 1.5, 4]),
                      Spacer(1, .5*cm)]

    if not has_data:
        from django.contrib import messages
        from django.shortcuts import redirect
        messages.warning(request, 'Nenhum dado encontrado para os filtros selecionados.')
        return redirect(request.META.get('HTTP_REFERER', '/relatorios/'))

    doc.build(story)
    resp = _pdf_resp(f'financeiro_{de}_{ate}.pdf')
    resp.write(buf.getvalue())
    return resp


@login_required
@requer_empresa
def financeiro_excel(request):
    de, ate  = _parse_periodo(request)
    subtipo  = request.GET.get('subtipo', 'despesas')
    centro   = request.GET.get('centro', '')
    subgrupo = request.GET.get('subgrupo', '')
    forma    = request.GET.get('forma', '')

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    label = f'{de.strftime("%d/%m/%Y")} a {ate.strftime("%d/%m/%Y")}'
    wb = Workbook()
    wb.remove(wb.active)

    def _sheet(ws, headers, rows, totais=None, min_widths=None):
        ws.append([ws.title])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws['A1'].font = Font(bold=True, size=13)
        ws.append([])
        ws.append(headers)
        hfill = PatternFill('solid', fgColor='1e1b4b')
        for c in ws[ws.max_row]:
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = hfill
            c.alignment = Alignment(horizontal='center')
        for i, row in enumerate(rows):
            ws.append(row)
            if i % 2 == 1:
                for c in ws[ws.max_row]:
                    c.fill = PatternFill('solid', fgColor='F8FAFC')
        if totais:
            ws.append(totais)
            for c in ws[ws.max_row]:
                c.font = Font(bold=True)
                c.fill = PatternFill('solid', fgColor='E0E7FF')
        for i, col in enumerate(ws.columns):
            try:
                from openpyxl.utils import get_column_letter
                letter = get_column_letter(i + 1)
                ml = max((len(str(c.value or '')) for c in col if hasattr(c, 'value') and not hasattr(c, 'parent') or True), default=10)
                # ignora MergedCell (sem atributo 'value' acessível) usando valor da célula diretamente
                vals = []
                for c in col:
                    try:
                        vals.append(len(str(c.value or '')))
                    except Exception:
                        pass
                ml = max(vals) if vals else 10
                w = min(ml + 4, 60)
                if min_widths and i < len(min_widths):
                    w = max(w, min_widths[i])
                ws.column_dimensions[letter].width = w
            except Exception:
                pass

    if subtipo in ('despesas', 'ambos'):
        qs = Despesa.objects.filter(empresa=request.empresa, data__range=(de, ate))
        if centro:   qs = qs.filter(centro_custo=centro)
        if subgrupo: qs = qs.filter(subgrupo=subgrupo)
        if forma:    qs = qs.filter(forma_pagamento=forma)
    has_data = False
    if subtipo in ('despesas', 'ambos'):
        qs = Despesa.objects.filter(empresa=request.empresa, data__range=(de, ate))
        if centro:   qs = qs.filter(centro_custo=centro)
        if subgrupo: qs = qs.filter(subgrupo=subgrupo)
        if forma:    qs = qs.filter(forma_pagamento=forma)
        if qs.exists():
            has_data = True
            rows  = [[_fmt_data(d.data), d.centro_custo, d.subgrupo, d.descricao,
                      d.forma_pagamento, d.get_situacao_display(), _fmt_valor(d.valor)] for d in qs]
            total = qs.aggregate(v=Sum('valor'))['v'] or 0
            ws = wb.create_sheet(f'Despesas')
            # Data, Centro, SubGrupo, Descrição, FormaPag, Situação, Valor
            _sheet(ws, ['Data','Centro','SubGrupo','Descrição','Forma Pag.','Situação','Valor (R$)'],
                   rows, ['','','','','','TOTAL', _fmt_valor(total)],
                   min_widths=[12, 25, 25, 45, 22, 12, 16])

    if subtipo in ('repasses', 'ambos'):
        qs = Repasse.objects.filter(empresa=request.empresa, data__range=(de, ate))
        if qs.exists():
            has_data = True
            rows  = [[_fmt_data(r.data), r.get_tipo_display(), r.origem, r.destino,
                      _fmt_valor(r.valor), r.descricao] for r in qs]
            total = qs.aggregate(v=Sum('valor'))['v'] or 0
            ws = wb.create_sheet('Repasses')
            # Data, Tipo, Origem, Destino, Valor, Descrição
            _sheet(ws, ['Data','Tipo','Origem','Destino','Valor (R$)','Descrição'],
                   rows, ['','','','TOTAL', _fmt_valor(total), ''],
                   min_widths=[12, 14, 25, 25, 16, 45])

    if not has_data:
        from django.contrib import messages
        from django.shortcuts import redirect
        messages.warning(request, 'Nenhum dado encontrado para os filtros selecionados.')
        return redirect(request.META.get('HTTP_REFERER', '/relatorios/'))

    buf = io.BytesIO()
    wb.save(buf)
    resp = _excel_resp(f'financeiro_{de}_{ate}.xlsx')
    resp.write(buf.getvalue())
    return resp


# ── FROTAS ────────────────────────────────────────────────────────────────────

def _frotas_qs(request, de, ate):
    """Retorna querysets filtrados para cada entidade de frota."""
    empresa    = request.empresa
    veiculo_pk = request.GET.get('veiculo') or None

    ab_qs  = Abastecimento.objects.filter(veiculo__empresa=empresa, data__range=(de, ate)).select_related('veiculo')
    man_qs = Manutencao.objects.filter(veiculo__empresa=empresa, data__range=(de, ate)).select_related('veiculo')
    via_qs = Viagem.objects.filter(veiculo__empresa=empresa, saida__date__range=(de, ate)).select_related('veiculo', 'motorista')
    vei_qs = Veiculo.objects.filter(empresa=empresa)

    if veiculo_pk:
        ab_qs  = ab_qs.filter(veiculo_id=veiculo_pk)
        man_qs = man_qs.filter(veiculo_id=veiculo_pk)
        via_qs = via_qs.filter(veiculo_id=veiculo_pk)

    return vei_qs, ab_qs, man_qs, via_qs


@login_required
@requer_empresa
def frotas_pdf(request):
    from django.contrib import messages
    from django.shortcuts import redirect
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    de, ate  = _parse_periodo(request)
    subtipo  = request.GET.get('subtipo', 'todos')
    label    = f'{de.strftime("%d/%m/%Y")} a {ate.strftime("%d/%m/%Y")}'
    vei_qs, ab_qs, man_qs, via_qs = _frotas_qs(request, de, ate)

    buf    = io.BytesIO()
    page_w = landscape(A4)[0] - 3*cm
    doc    = SimpleDocTemplate(buf, pagesize=landscape(A4),
                               leftMargin=1.5*cm, rightMargin=1.5*cm,
                               topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story  = []

    logo_path = request.empresa.logo.path if request.empresa.logo else None
    if logo_path:
        from reportlab.platypus import Image as RLImage
        try:
            story.append(RLImage(logo_path, height=1.2*cm, width=3.5*cm, kind='proportional'))
            story.append(Spacer(1, .2*cm))
        except Exception:
            pass
    story += [Paragraph(f'Relatório de Frotas — {label}', styles['Title']), Spacer(1, .4*cm)]

    def _tbl(headers, rows, totais=None, col_parts=None):
        data = [headers] + rows + ([totais] if totais else [])
        parts = col_parts or ([1] * len(headers))
        w = [page_w * (p / sum(parts)) for p in parts]
        t = Table(data, colWidths=w, repeatRows=1)
        s = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e1b4b')),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1,-1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
            ('GRID', (0, 0), (-1,-1), .3, colors.HexColor('#e2e8f0')),
            ('TOPPADDING',    (0, 0), (-1,-1), 4),
            ('BOTTOMPADDING', (0, 0), (-1,-1), 4),
        ]
        if totais:
            s += [('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#e0e7ff')),
                  ('FONTNAME',   (0,-1), (-1,-1), 'Helvetica-Bold')]
        t.setStyle(TableStyle(s))
        return t

    has_data = False

    # ── Veículos ──
    if subtipo in ('veiculos', 'todos'):
        if vei_qs.exists():
            has_data = True
            rows = [[v.placa, v.marca, v.modelo, str(v.ano),
                     f'{v.km_atual:,} km'.replace(',', '.'),
                     v.get_status_display()] for v in vei_qs]
            story += [Paragraph('Veículos', styles['Heading2']), Spacer(1, .2*cm),
                      _tbl(['Placa', 'Marca', 'Modelo', 'Ano', 'KM Atual', 'Status'],
                           rows, col_parts=[1.5, 1.5, 2, 1, 1.5, 1.5]),
                      Spacer(1, .6*cm)]

    # ── Abastecimentos ──
    if subtipo in ('abastecimentos', 'todos'):
        if ab_qs.exists():
            has_data = True
            total_val   = ab_qs.aggregate(v=Sum('valor_total'))['v'] or 0
            total_litros = ab_qs.aggregate(l=Sum('litros'))['l'] or 0
            rows = [[_fmt_data(a.data), a.veiculo.placa,
                     f'{a.veiculo.marca} {a.veiculo.modelo}',
                     f'{a.km_atual:,} km'.replace(',', '.'),
                     f'{a.litros} L', _fmt_valor(a.valor_total),
                     a.posto or '—'] for a in ab_qs]
            story += [Paragraph('Abastecimentos', styles['Heading2']), Spacer(1, .2*cm),
                      _tbl(['Data', 'Veículo', 'Marca/Modelo', 'KM', 'Litros', 'Valor', 'Posto'],
                           rows,
                           ['', '', '', '', f'Total: {total_litros} L', _fmt_valor(total_val), ''],
                           col_parts=[1, 1.2, 2, 1.3, 1.2, 1.5, 2.3]),
                      Spacer(1, .6*cm)]

    # ── Manutenções ──
    if subtipo in ('manutencoes', 'todos'):
        if man_qs.exists():
            has_data = True
            total_val = man_qs.filter(valor__isnull=False).aggregate(v=Sum('valor'))['v'] or 0
            rows = [[_fmt_data(m.data), m.veiculo.placa,
                     f'{m.veiculo.marca} {m.veiculo.modelo}',
                     m.get_tipo_display(), m.descricao[:50],
                     m.get_status_display(),
                     _fmt_valor(m.valor) if m.valor else '—',
                     m.oficina or '—'] for m in man_qs]
            story += [Paragraph('Manutenções', styles['Heading2']), Spacer(1, .2*cm),
                      _tbl(['Data', 'Veículo', 'Marca/Modelo', 'Tipo', 'Descrição', 'Status', 'Valor', 'Oficina'],
                           rows,
                           ['', '', '', '', '', 'TOTAL', _fmt_valor(total_val), ''],
                           col_parts=[1, 1.2, 2, 1.2, 3, 1.2, 1.5, 2]),
                      Spacer(1, .6*cm)]

    # ── Viagens ──
    if subtipo in ('viagens', 'todos'):
        if via_qs.exists():
            has_data = True
            rows = [[_fmt_data(v.saida), v.veiculo.placa,
                     f'{v.veiculo.marca} {v.veiculo.modelo}',
                     v.motorista.nome if v.motorista else '—',
                     v.origem, v.destino,
                     f'{v.km_rodado:,} km'.replace(',', '.') if v.km_rodado else '—',
                     v.get_status_display()] for v in via_qs]
            story += [Paragraph('Viagens', styles['Heading2']), Spacer(1, .2*cm),
                      _tbl(['Data', 'Veículo', 'Marca/Modelo', 'Motorista', 'Origem', 'Destino', 'KM Rod.', 'Status'],
                           rows, col_parts=[1, 1.2, 2, 2, 1.8, 1.8, 1.2, 1.2]),
                      Spacer(1, .6*cm)]

    if not has_data:
        messages.warning(request, 'Nenhum dado encontrado para os filtros selecionados.')
        return redirect(request.META.get('HTTP_REFERER', '/relatorios/'))

    doc.build(story)
    resp = _pdf_resp(f'frotas_{de}_{ate}.pdf')
    resp.write(buf.getvalue())
    return resp


@login_required
@requer_empresa
def frotas_excel(request):
    from django.contrib import messages
    from django.shortcuts import redirect
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    de, ate  = _parse_periodo(request)
    subtipo  = request.GET.get('subtipo', 'todos')
    label    = f'{de.strftime("%d/%m/%Y")} a {ate.strftime("%d/%m/%Y")}'
    vei_qs, ab_qs, man_qs, via_qs = _frotas_qs(request, de, ate)

    wb = Workbook()
    wb.remove(wb.active)

    def _sheet(ws, headers, rows, totais=None, min_widths=None):
        ws.append([ws.title])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws['A1'].font = Font(bold=True, size=13)
        ws.append([])
        ws.append(headers)
        hfill = PatternFill('solid', fgColor='1e1b4b')
        for c in ws[ws.max_row]:
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = hfill
            c.alignment = Alignment(horizontal='center')
        for i, row in enumerate(rows):
            ws.append(row)
            if i % 2 == 1:
                for c in ws[ws.max_row]:
                    c.fill = PatternFill('solid', fgColor='F8FAFC')
        if totais:
            ws.append(totais)
            for c in ws[ws.max_row]:
                c.font = Font(bold=True)
                c.fill = PatternFill('solid', fgColor='E0E7FF')
        for i, col in enumerate(ws.columns):
            try:
                letter = get_column_letter(i + 1)
                vals = []
                for c in col:
                    try:
                        vals.append(len(str(c.value or '')))
                    except Exception:
                        pass
                w = min(max(vals, default=10) + 4, 60)
                if min_widths and i < len(min_widths):
                    w = max(w, min_widths[i])
                ws.column_dimensions[letter].width = w
            except Exception:
                pass

    has_data = False

    if subtipo in ('veiculos', 'todos') and vei_qs.exists():
        has_data = True
        rows = [[v.placa, v.marca, v.modelo, v.ano, v.km_atual, v.get_status_display()]
                for v in vei_qs]
        ws = wb.create_sheet('Veículos')
        _sheet(ws, ['Placa', 'Marca', 'Modelo', 'Ano', 'KM Atual', 'Status'],
               rows, min_widths=[12, 16, 20, 6, 12, 14])

    if subtipo in ('abastecimentos', 'todos') and ab_qs.exists():
        has_data = True
        total_val    = ab_qs.aggregate(v=Sum('valor_total'))['v'] or 0
        total_litros = ab_qs.aggregate(l=Sum('litros'))['l'] or 0
        rows = [[_fmt_data(a.data), a.veiculo.placa, f'{a.veiculo.marca} {a.veiculo.modelo}',
                 a.km_atual, float(a.litros), float(a.valor_total), a.posto or ''] for a in ab_qs]
        ws = wb.create_sheet('Abastecimentos')
        _sheet(ws, ['Data', 'Veículo', 'Marca/Modelo', 'KM', 'Litros', 'Valor (R$)', 'Posto'],
               rows,
               ['', '', '', '', f'Total: {total_litros} L', _fmt_valor(total_val), ''],
               min_widths=[12, 12, 22, 12, 10, 16, 30])

    if subtipo in ('manutencoes', 'todos') and man_qs.exists():
        has_data = True
        total_val = man_qs.filter(valor__isnull=False).aggregate(v=Sum('valor'))['v'] or 0
        rows = [[_fmt_data(m.data), m.veiculo.placa, f'{m.veiculo.marca} {m.veiculo.modelo}',
                 m.get_tipo_display(), m.descricao, m.get_status_display(),
                 float(m.valor) if m.valor else '', m.oficina or ''] for m in man_qs]
        ws = wb.create_sheet('Manutenções')
        _sheet(ws, ['Data', 'Veículo', 'Marca/Modelo', 'Tipo', 'Descrição', 'Status', 'Valor (R$)', 'Oficina'],
               rows,
               ['', '', '', '', '', 'TOTAL', _fmt_valor(total_val), ''],
               min_widths=[12, 12, 22, 14, 40, 12, 16, 25])

    if subtipo in ('viagens', 'todos') and via_qs.exists():
        has_data = True
        rows = [[_fmt_data(v.saida), v.veiculo.placa, f'{v.veiculo.marca} {v.veiculo.modelo}',
                 v.motorista.nome if v.motorista else '',
                 v.origem, v.destino,
                 v.km_rodado or '', v.get_status_display()] for v in via_qs]
        ws = wb.create_sheet('Viagens')
        _sheet(ws, ['Data', 'Veículo', 'Marca/Modelo', 'Motorista', 'Origem', 'Destino', 'KM Rod.', 'Status'],
               rows, min_widths=[12, 12, 22, 25, 20, 20, 10, 14])

    if not has_data:
        messages.warning(request, 'Nenhum dado encontrado para os filtros selecionados.')
        return redirect(request.META.get('HTTP_REFERER', '/relatorios/'))

    buf = io.BytesIO()
    wb.save(buf)
    resp = _excel_resp(f'frotas_{de}_{ate}.xlsx')
    resp.write(buf.getvalue())
    return resp
