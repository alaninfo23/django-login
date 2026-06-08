from datetime import date
from decimal import Decimal
from datetime import date

from django.test import TestCase
from django.urls import reverse
from django.contrib.auth.models import User
from django.utils.text import slugify

from empresa.models import Empresa, MembroEmpresa
from assets.models import Asset
from financeiro.models import CentroCusto, Despesa, FormaPagamento, Repasse, SubGrupo
from frota.models import Veiculo, Abastecimento, Manutencao, Motorista, Viagem
from relatorios.views import _fmt_valor, _fmt_data, _parse_periodo


# ── helpers ────────────────────────────────────────────────────────────────────

def make_user(username='user', password='senha123'):
    return User.objects.create_user(username=username, password=password)


def make_empresa(nome='Empresa Teste'):
    return Empresa.objects.create(nome=nome, slug=slugify(nome))


def make_membro(user, empresa, perfil='admin'):
    return MembroEmpresa.objects.create(empresa=empresa, user=user, perfil=perfil)


def make_asset(user, empresa, **kwargs):
    defaults = dict(
        name='Notebook', asset_type=Asset.AssetType.IT,
        acquisition_value=Decimal('3000.00'),
        purchase_date=date(2026, 6, 1),
        status=Asset.Status.ACTIVE,
    )
    defaults.update(kwargs)
    return Asset.objects.create(user=user, empresa=empresa, **defaults)


def make_despesa(user, empresa, **kwargs):
    defaults = dict(
        data=date(2026, 6, 1), centro_custo='TI', subgrupo='Software',
        descricao='Licença', valor=Decimal('500.00'),
        forma_pagamento='Pix', situacao='pendente',
    )
    defaults.update(kwargs)
    return Despesa.objects.create(user=user, empresa=empresa, **defaults)


def make_repasse(user, empresa, **kwargs):
    defaults = dict(
        data=date(2026, 6, 1), origem='Matriz', destino='Filial',
        valor=Decimal('1000.00'), tipo=Repasse.Tipo.APORTE,
    )
    defaults.update(kwargs)
    return Repasse.objects.create(user=user, empresa=empresa, **defaults)


# ── Funções utilitárias (unitários puros) ──────────────────────────────────────

class FmtValorTests(TestCase):

    def test_formata_valor_simples(self):
        """Deve formatar Decimal para padrão brasileiro."""
        self.assertEqual(_fmt_valor(Decimal('1000.00')), 'R$ 1.000,00')

    def test_formata_valor_grande(self):
        """Deve formatar valores com múltiplos separadores de milhar."""
        self.assertEqual(_fmt_valor(Decimal('1234567.89')), 'R$ 1.234.567,89')

    def test_formata_zero(self):
        """Zero deve ser formatado como R$ 0,00."""
        self.assertEqual(_fmt_valor(Decimal('0')), 'R$ 0,00')

    def test_formata_float(self):
        """Deve aceitar float além de Decimal."""
        self.assertEqual(_fmt_valor(1500.50), 'R$ 1.500,50')


class FmtDataTests(TestCase):

    def test_formata_objeto_date(self):
        """Deve formatar objeto date para dd/mm/yyyy."""
        self.assertEqual(_fmt_data(date(2026, 6, 7)), '07/06/2026')

    def test_formata_string_iso(self):
        """Deve formatar string ISO para dd/mm/yyyy."""
        self.assertEqual(_fmt_data('2026-06-07'), '07/06/2026')

    def test_retorna_vazio_para_none(self):
        """None deve retornar string vazia."""
        self.assertEqual(_fmt_data(None), '')

    def test_retorna_vazio_para_string_vazia(self):
        """String vazia deve retornar string vazia."""
        self.assertEqual(_fmt_data(''), '')


class ParsePeriodoTests(TestCase):

    def test_parse_datas_validas(self):
        """Deve parsear datas válidas corretamente."""
        req = self.client.get('/relatorios/?de=2026-01-01&ate=2026-06-30').wsgi_request
        de, ate = _parse_periodo(req)
        self.assertEqual(de, date(2026, 1, 1))
        self.assertEqual(ate, date(2026, 6, 30))

    def test_fallback_para_mes_atual_com_datas_invalidas(self):
        """Datas inválidas devem usar o mês atual como fallback."""
        req = self.client.get('/relatorios/?de=invalido&ate=invalido').wsgi_request
        de, ate = _parse_periodo(req)
        hoje = date.today()
        self.assertEqual(de, hoje.replace(day=1))
        self.assertEqual(ate, hoje)


# ── Views: autenticação ────────────────────────────────────────────────────────

class RelatoriosAuthTests(TestCase):

    PROTECTED = [
        'relatorios', 'rel_patrimonial_pdf', 'rel_patrimonial_excel',
        'rel_financeiro_pdf', 'rel_financeiro_excel',
    ]

    def test_views_redirecionam_sem_login(self):
        """Todas as views de relatórios devem redirecionar sem autenticação."""
        for name in self.PROTECTED:
            resp = self.client.get(reverse(name))
            self.assertEqual(resp.status_code, 302, msg=f'{name} deveria redirecionar')
            self.assertIn('/login', resp['Location'])


# ── View: central ──────────────────────────────────────────────────────────────

class CentralViewTests(TestCase):

    def setUp(self):
        self.user = make_user()
        self.empresa = make_empresa()
        make_membro(self.user, self.empresa)
        self.client.force_login(self.user)

    def test_retorna_200(self):
        """Página central de relatórios deve retornar 200."""
        self.assertEqual(self.client.get(reverse('relatorios')).status_code, 200)

    def test_aba_padrao_e_patrimonial(self):
        """Módulo padrão deve ser patrimonial."""
        resp = self.client.get(reverse('relatorios'))
        self.assertEqual(resp.context['modulo'], 'patrimonial')

    def test_aba_financeiro_via_querystring(self):
        """Deve aceitar módulo via querystring."""
        resp = self.client.get(reverse('relatorios') + '?modulo=financeiro')
        self.assertEqual(resp.context['modulo'], 'financeiro')

    def test_contexto_contem_tipos_e_filtros(self):
        """Contexto deve conter asset_types, centros, subgrupos e formas."""
        resp = self.client.get(reverse('relatorios'))
        for key in ('asset_types', 'asset_statuses', 'centros', 'subgrupos', 'formas'):
            self.assertIn(key, resp.context)


# ── View: patrimonial_pdf ──────────────────────────────────────────────────────

class PatrimonialPdfTests(TestCase):

    def setUp(self):
        self.user = make_user()
        self.empresa = make_empresa()
        make_membro(self.user, self.empresa)
        self.client.force_login(self.user)
        self.url = reverse('rel_patrimonial_pdf')

    def test_gera_pdf_com_dados(self):
        """Deve gerar PDF com content-type correto quando há dados."""
        make_asset(self.user, self.empresa)
        resp = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/pdf')

    def test_redireciona_sem_dados(self):
        """Deve redirecionar com aviso quando não há dados no período."""
        resp = self.client.get(self.url + '?de=2020-01-01&ate=2020-01-31')
        self.assertEqual(resp.status_code, 302)

    def test_filtro_por_tipo(self):
        """Filtro por tipo deve retornar apenas assets do tipo selecionado."""
        make_asset(self.user, self.empresa, asset_type='ti')
        make_asset(self.user, self.empresa, asset_type='carro', name='Fusca')
        # com filtro só TI: carro fica fora → PDF gerado com 1 item
        resp = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30&subtipo=ti')
        self.assertEqual(resp.status_code, 200)
        # sem filtro: ambos presentes → também gera PDF
        resp2 = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30')
        self.assertEqual(resp2.status_code, 200)
        # PDF com filtro deve ser menor que sem filtro (1 linha vs 2)
        self.assertLess(len(resp.content), len(resp2.content))

    def test_filename_no_header(self):
        """Nome do arquivo no header deve conter 'patrimonial'."""
        make_asset(self.user, self.empresa)
        resp = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30')
        self.assertIn('patrimonial', resp['Content-Disposition'])


# ── View: patrimonial_excel ────────────────────────────────────────────────────

class PatrimonialExcelTests(TestCase):

    def setUp(self):
        self.user = make_user()
        self.empresa = make_empresa()
        make_membro(self.user, self.empresa)
        self.client.force_login(self.user)
        self.url = reverse('rel_patrimonial_excel')

    def test_gera_excel_com_dados(self):
        """Deve gerar Excel com content-type correto quando há dados."""
        make_asset(self.user, self.empresa)
        resp = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30')
        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheet', resp['Content-Type'])

    def test_redireciona_sem_dados(self):
        """Deve redirecionar com aviso quando não há dados no período."""
        resp = self.client.get(self.url + '?de=2020-01-01&ate=2020-01-31')
        self.assertEqual(resp.status_code, 302)

    def test_excel_contem_nome_do_bem(self):
        """Arquivo Excel deve conter o nome do ativo cadastrado."""
        make_asset(self.user, self.empresa, name='MacBook Pro')
        resp = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30')
        self.assertEqual(resp.status_code, 200)
        # lê o Excel e verifica o conteúdo das células
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        valores = [str(c.value or '') for row in ws.iter_rows() for c in row]
        self.assertIn('MacBook Pro', valores)


# ── View: financeiro_pdf ───────────────────────────────────────────────────────

class FinanceiroPdfTests(TestCase):

    def setUp(self):
        self.user = make_user()
        self.empresa = make_empresa()
        make_membro(self.user, self.empresa)
        self.client.force_login(self.user)
        self.url = reverse('rel_financeiro_pdf')

    def test_gera_pdf_despesas(self):
        """Deve gerar PDF de despesas com content-type correto."""
        make_despesa(self.user, self.empresa)
        resp = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30&subtipo=despesas')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/pdf')

    def test_gera_pdf_repasses(self):
        """Deve gerar PDF de repasses com content-type correto."""
        make_repasse(self.user, self.empresa)
        resp = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30&subtipo=repasses')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/pdf')

    def test_gera_pdf_ambos(self):
        """Deve gerar PDF com despesas e repasses juntos."""
        make_despesa(self.user, self.empresa)
        make_repasse(self.user, self.empresa)
        resp = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30&subtipo=ambos')
        self.assertEqual(resp.status_code, 200)

    def test_redireciona_sem_dados(self):
        """Deve redirecionar quando não há dados no período."""
        resp = self.client.get(self.url + '?de=2020-01-01&ate=2020-01-31&subtipo=despesas')
        self.assertEqual(resp.status_code, 302)


# ── View: financeiro_excel ─────────────────────────────────────────────────────

class FinanceiroExcelTests(TestCase):

    def setUp(self):
        self.user = make_user()
        self.empresa = make_empresa()
        make_membro(self.user, self.empresa)
        self.client.force_login(self.user)
        self.url = reverse('rel_financeiro_excel')

    def test_gera_excel_despesas(self):
        """Deve gerar Excel de despesas com content-type correto."""
        make_despesa(self.user, self.empresa)
        resp = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30&subtipo=despesas')
        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheet', resp['Content-Type'])

    def test_gera_excel_repasses(self):
        """Deve gerar Excel de repasses."""
        make_repasse(self.user, self.empresa)
        resp = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30&subtipo=repasses')
        self.assertEqual(resp.status_code, 200)

    def test_redireciona_sem_dados(self):
        """Deve redirecionar quando não há dados no período."""
        resp = self.client.get(self.url + '?de=2020-01-01&ate=2020-01-31&subtipo=despesas')
        self.assertEqual(resp.status_code, 302)

    def test_excel_contem_descricao_despesa(self):
        """Excel deve conter a descrição da despesa cadastrada."""
        make_despesa(self.user, self.empresa, descricao='Assinatura Adobe')
        resp = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30&subtipo=despesas')
        self.assertEqual(resp.status_code, 200)
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        valores = [str(c.value or '') for row in ws.iter_rows() for c in row]
        self.assertIn('Assinatura Adobe', valores)


# ── helpers de frota ──────────────────────────────────────────────────────────

def make_veiculo(empresa, placa='ABC-1234', km_atual=1000):
    return Veiculo.objects.create(
        empresa=empresa, placa=placa, modelo='Gol', marca='VW',
        ano=2020, capacidade_carga='1.5', km_atual=km_atual,
    )


def make_abastecimento_rel(veiculo):
    return Abastecimento.objects.create(
        veiculo=veiculo, data=date(2026, 6, 1),
        km_atual=1200, litros='50.00', valor_total='250.00',
    )


def make_manutencao_rel(veiculo):
    return Manutencao.objects.create(
        veiculo=veiculo, tipo=Manutencao.Tipo.PREVENTIVA,
        descricao='Troca de óleo', status=Manutencao.Status.REALIZADA,
        data=date(2026, 6, 1), valor='350.00',
    )


def make_viagem_rel(veiculo):
    from datetime import datetime
    return Viagem.objects.create(
        veiculo=veiculo, origem='SP', destino='RJ',
        saida=datetime(2026, 6, 1, 8, 0),
        km_inicial=1000, km_final=1400,
        status=Viagem.Status.CONCLUIDA,
    )


# ── View: frotas_pdf ──────────────────────────────────────────────────────────

class FrotasPdfTests(TestCase):

    def setUp(self):
        self.user = make_user()
        self.empresa = make_empresa()
        make_membro(self.user, self.empresa)
        self.client.force_login(self.user)
        self.veiculo = make_veiculo(self.empresa)
        self.url = reverse('rel_frotas_pdf')

    def test_gera_pdf_veiculos(self):
        resp = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30&subtipo=veiculos')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/pdf')

    def test_gera_pdf_abastecimentos(self):
        make_abastecimento_rel(self.veiculo)
        resp = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30&subtipo=abastecimentos')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/pdf')

    def test_gera_pdf_manutencoes(self):
        make_manutencao_rel(self.veiculo)
        resp = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30&subtipo=manutencoes')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/pdf')

    def test_gera_pdf_viagens(self):
        make_viagem_rel(self.veiculo)
        resp = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30&subtipo=viagens')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/pdf')

    def test_gera_pdf_todos(self):
        make_abastecimento_rel(self.veiculo)
        resp = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30&subtipo=todos')
        self.assertEqual(resp.status_code, 200)

    def test_redireciona_sem_dados_abastecimentos(self):
        """Sem abastecimentos no período deve redirecionar."""
        resp = self.client.get(self.url + '?de=2020-01-01&ate=2020-01-31&subtipo=abastecimentos')
        self.assertEqual(resp.status_code, 302)

    def test_filtro_por_veiculo(self):
        """Com veiculo_pk específico só inclui registros daquele veículo."""
        make_abastecimento_rel(self.veiculo)
        outro = make_veiculo(self.empresa, placa='ZZZ-9999')
        resp = self.client.get(
            self.url + f'?de=2026-06-01&ate=2026-06-30&subtipo=abastecimentos&veiculo={outro.pk}'
        )
        # sem abastecimento para esse veículo — redireciona
        self.assertEqual(resp.status_code, 302)

    def test_filename_contem_frotas(self):
        resp = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30&subtipo=veiculos')
        self.assertIn('frotas', resp['Content-Disposition'])

    def test_nao_acessa_sem_login(self):
        self.client.logout()
        resp = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30')
        self.assertEqual(resp.status_code, 302)


# ── View: frotas_excel ────────────────────────────────────────────────────────

class FrotasExcelTests(TestCase):

    def setUp(self):
        self.user = make_user()
        self.empresa = make_empresa()
        make_membro(self.user, self.empresa)
        self.client.force_login(self.user)
        self.veiculo = make_veiculo(self.empresa)
        self.url = reverse('rel_frotas_excel')

    def test_gera_excel_veiculos(self):
        resp = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30&subtipo=veiculos')
        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheet', resp['Content-Type'])

    def test_gera_excel_abastecimentos(self):
        make_abastecimento_rel(self.veiculo)
        resp = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30&subtipo=abastecimentos')
        self.assertEqual(resp.status_code, 200)
        self.assertIn('spreadsheet', resp['Content-Type'])

    def test_gera_excel_manutencoes(self):
        make_manutencao_rel(self.veiculo)
        resp = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30&subtipo=manutencoes')
        self.assertEqual(resp.status_code, 200)

    def test_gera_excel_viagens(self):
        make_viagem_rel(self.veiculo)
        resp = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30&subtipo=viagens')
        self.assertEqual(resp.status_code, 200)

    def test_gera_excel_todos(self):
        make_abastecimento_rel(self.veiculo)
        resp = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30&subtipo=todos')
        self.assertEqual(resp.status_code, 200)

    def test_redireciona_sem_dados(self):
        resp = self.client.get(self.url + '?de=2020-01-01&ate=2020-01-31&subtipo=abastecimentos')
        self.assertEqual(resp.status_code, 302)

    def test_excel_abastecimentos_contem_placa(self):
        """Excel de abastecimentos deve conter a placa do veículo."""
        make_abastecimento_rel(self.veiculo)
        resp = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30&subtipo=abastecimentos')
        self.assertEqual(resp.status_code, 200)
        import io as _io
        from openpyxl import load_workbook
        wb = load_workbook(_io.BytesIO(resp.content))
        ws = wb.active
        valores = [str(c.value or '') for row in ws.iter_rows() for c in row]
        self.assertIn('ABC-1234', valores)

    def test_excel_todos_gera_multiplas_abas(self):
        """subtipo=todos deve gerar abas para cada entidade com dados."""
        make_abastecimento_rel(self.veiculo)
        make_manutencao_rel(self.veiculo)
        resp = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30&subtipo=todos')
        self.assertEqual(resp.status_code, 200)
        import io as _io
        from openpyxl import load_workbook
        wb = load_workbook(_io.BytesIO(resp.content))
        # deve ter aba Veículos, Abastecimentos e Manutenções
        self.assertIn('Veículos', wb.sheetnames)
        self.assertIn('Abastecimentos', wb.sheetnames)
        self.assertIn('Manutenções', wb.sheetnames)

    def test_isolamento_entre_empresas(self):
        """Relatório não deve incluir dados de outra empresa."""
        outra = make_empresa('Outra')
        v2 = make_veiculo(outra, placa='OUT-0001')
        make_abastecimento_rel(v2)
        resp = self.client.get(self.url + '?de=2026-06-01&ate=2026-06-30&subtipo=abastecimentos')
        # empresa do usuário logado não tem abastecimentos → redireciona
        self.assertEqual(resp.status_code, 302)
